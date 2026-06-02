"""Génération du récap Excel : Synthèse, Par véhicule, Excès vitesse, Activité."""
import io
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage

NAVY = "003366"
BLUE = "2E6DA4"
LIGHT = "EAF0F6"
WHITE = "FFFFFF"
RED = "C0392B"

ASSETS = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _title_band(ws, meta, kpis, ncols=8, title="RÉCAP TRAJETS"):
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=1, column=1)
    c.value = (f"{title} — {kpis['client']}\n"
               f"Période : {meta.get('period_start','?')} → {meta.get('period_end','?')}  |  "
               f"Comafrique Technologies — Support Tracking")
    c.font = Font(name="Arial", size=12, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    for col in range(1, ncols + 1):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor=NAVY)
    # logos
    for fname, anchor in [("logo_comafrique.png", None), ("logo_client_placeholder.png", None)]:
        pass  # logos ajoutés ci-dessous via fonction dédiée


def _add_logos(ws, ncols):
    lp = os.path.join(ASSETS, "logo_comafrique.png")
    rp = os.path.join(ASSETS, "logo_client_placeholder.png")
    try:
        if os.path.exists(lp):
            img = XLImage(lp)
            img.height, img.width = 44, 114
            ws.add_image(img, "A1")
        if os.path.exists(rp):
            img2 = XLImage(rp)
            img2.height, img2.width = 44, 114
            ws.add_image(img2, f"{get_column_letter(max(ncols-1,2))}1")
    except Exception:
        pass


def _header_row(ws, row, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_excel(result, meta) -> bytes:
    k = result["kpis"]
    wb = Workbook()

    # ============ Feuille SYNTHÈSE ============
    ws = wb.active
    ws.title = "Synthèse"
    _title_band(ws, meta, k, ncols=4)
    _add_logos(ws, 4)
    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16

    rows = [
        ("Indicateur", "Valeur"),
        ("Nombre de trajets", k["n_trips"]),
        ("Véhicules actifs", k["n_vehicles"]),
        ("Jours d'activité", k["n_days"]),
        ("Distance totale (km)", k["total_distance"]),
        ("Durée de conduite (h)", k["total_duration_h"]),
        ("Distance moy. / trajet (km)", k["avg_distance"]),
        ("Durée moy. / trajet (min)", k["avg_duration_min"]),
        ("Trajets moy. / véhicule", k["avg_trips_per_vehicle"]),
        ("Km moy. / véhicule", k["avg_km_per_vehicle"]),
        ("Vitesse max flotte (km/h)", k["max_speed_fleet"]),
        ("Vitesse max moyenne (km/h)", k["avg_max_speed"]),
        (f"Excès vitesse (> {k['speed_threshold']} km/h)", k["n_speeding"]),
        ("Part excès vitesse (%)", k["pct_speeding"]),
        (f"Trajets de nuit ({k['night_window']})", k["n_night"]),
        ("Part trajets nuit (%)", k["pct_night"]),
    ]
    start = 5
    _header_row(ws, start, rows[0])
    for i, (lbl, val) in enumerate(rows[1:], start=start + 1):
        a = ws.cell(row=i, column=1, value=lbl)
        b = ws.cell(row=i, column=2, value=val)
        a.font = Font(name="Arial", size=10)
        b.font = Font(name="Arial", size=10, bold=True, color=NAVY)
        a.border = BORDER; b.border = BORDER
        a.fill = PatternFill("solid", fgColor=LIGHT if i % 2 else WHITE)
        b.fill = PatternFill("solid", fgColor=LIGHT if i % 2 else WHITE)
        b.alignment = Alignment(horizontal="right")
    _autosize(ws, {"A": 34, "B": 16})

    # ============ Feuille PAR VÉHICULE ============
    wsv = wb.create_sheet("Par véhicule")
    bv = result["by_vehicle"]
    _title_band(wsv, meta, k, ncols=len(bv.columns))
    _add_logos(wsv, len(bv.columns))
    hr = 5
    _header_row(wsv, hr, list(bv.columns))
    for i, (_, r) in enumerate(bv.iterrows(), start=hr + 1):
        for j, v in enumerate(r, start=1):
            c = wsv.cell(row=i, column=j, value=v)
            c.font = Font(name="Arial", size=9)
            c.border = BORDER
            c.fill = PatternFill("solid", fgColor=LIGHT if i % 2 else WHITE)
            if j > 1:
                c.alignment = Alignment(horizontal="center")
            # surlignage excès
            if bv.columns[j-1] == "Excès vitesse" and isinstance(v, (int, float)) and v > 0:
                c.font = Font(name="Arial", size=9, bold=True, color=RED)
    # Ligne TOTAL avec formules
    last = hr + len(bv)
    tot = last + 1
    wsv.cell(row=tot, column=1, value="TOTAL / MOYENNE").font = Font(bold=True, color=WHITE)
    wsv.cell(row=tot, column=1).fill = PatternFill("solid", fgColor=NAVY)
    col_map = {c: idx + 1 for idx, c in enumerate(bv.columns)}
    def colL(name): return get_column_letter(col_map[name])
    formulas = {
        "Trajets": f"=SUM({colL('Trajets')}{hr+1}:{colL('Trajets')}{last})",
        "Distance (km)": f"=SUM({colL('Distance (km)')}{hr+1}:{colL('Distance (km)')}{last})",
        "Durée (h)": f"=SUM({colL('Durée (h)')}{hr+1}:{colL('Durée (h)')}{last})",
        "V. max (km/h)": f"=MAX({colL('V. max (km/h)')}{hr+1}:{colL('V. max (km/h)')}{last})",
        "V. moy (km/h)": f"=ROUND(AVERAGE({colL('V. moy (km/h)')}{hr+1}:{colL('V. moy (km/h)')}{last}),1)",
        "Excès vitesse": f"=SUM({colL('Excès vitesse')}{hr+1}:{colL('Excès vitesse')}{last})",
        "Trajets nuit": f"=SUM({colL('Trajets nuit')}{hr+1}:{colL('Trajets nuit')}{last})",
    }
    for name, f in formulas.items():
        c = wsv.cell(row=tot, column=col_map[name], value=f)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
    _autosize(wsv, {"A": 30, "B": 10, "C": 13, "D": 11, "E": 13, "F": 13, "G": 13, "H": 13})
    wsv.freeze_panes = f"A{hr+1}"

    # ============ Feuille EXCÈS VITESSE ============
    wse = wb.create_sheet("Excès vitesse")
    sd = result["speeding_detail"]
    _title_band(wse, meta, k, ncols=6)
    _add_logos(wse, 6)
    _header_row(wse, 5, list(sd.columns))
    for i, (_, r) in enumerate(sd.iterrows(), start=6):
        for j, v in enumerate(r, start=1):
            c = wse.cell(row=i, column=j, value=v)
            c.font = Font(name="Arial", size=8)
            c.border = BORDER
            c.fill = PatternFill("solid", fgColor=LIGHT if i % 2 else WHITE)
    _autosize(wse, {"A": 26, "B": 20, "C": 12, "D": 13, "E": 42, "F": 42})

    # ============ Feuille ACTIVITÉ JOUR ============
    wsd = wb.create_sheet("Activité journalière")
    bd = result["by_day"].copy()
    bd["date"] = bd["date"].dt.strftime("%d/%m/%Y")
    bd.columns = ["Date", "Trajets", "Distance (km)"]
    _title_band(wsd, meta, k, ncols=3)
    _add_logos(wsd, 3)
    _header_row(wsd, 5, list(bd.columns))
    for i, (_, r) in enumerate(bd.iterrows(), start=6):
        for j, v in enumerate(r, start=1):
            c = wsd.cell(row=i, column=j, value=v)
            c.font = Font(name="Arial", size=9)
            c.border = BORDER
            c.fill = PatternFill("solid", fgColor=LIGHT if i % 2 else WHITE)
            if j > 1:
                c.alignment = Alignment(horizontal="center")
    _autosize(wsd, {"A": 14, "B": 12, "C": 14})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _write_table(ws, df, header_row, start_data, num_cols_center=True):
    _header_row(ws, header_row, list(df.columns))
    for i, (_, r) in enumerate(df.iterrows(), start=start_data):
        for j, v in enumerate(r, start=1):
            if hasattr(v, "item"):
                try:
                    v = v.item()
                except Exception:
                    pass
            c = ws.cell(row=i, column=j, value=(None if pd.isna(v) else v))
            c.font = Font(name="Arial", size=9)
            c.border = BORDER
            c.fill = PatternFill("solid", fgColor=LIGHT if i % 2 else WHITE)
            if num_cols_center and j > 1:
                c.alignment = Alignment(horizontal="center")
    return start_data + len(df) - 1


def build_scorecard_excel(result, meta) -> bytes:
    """Récap Excel combiné : Synthèse, Scorecard, Comportement détail, Conso détail."""
    import pandas as pd  # local pour _write_table
    k = result["kpis"]
    wb = Workbook()

    # ---- Synthèse ----
    ws = wb.active
    ws.title = "Synthèse"
    _title_band(ws, meta, k, ncols=4, title="SCORECARD FLOTTE")
    _add_logos(ws, 4)
    rows = [
        ("Indicateur", "Valeur"),
        ("Véhicules (total)", k["n_vehicles"]),
        ("Véhicules notés Ecodrive", k["n_with_behavior"]),
        ("Véhicules avec données carburant", k["n_with_fuel"]),
        ("Conducteurs", k["n_drivers"]),
        ("Distance totale (km)", k["total_distance"]),
        ("Carburant total (L)", k["total_fuel"]),
        ("Score moyen flotte (%)", k["avg_score"]),
        ("Conso moyenne (L/100km)", k["avg_conso"] if k["avg_conso"] else "n/a"),
        ("Événements brusques (total)", k["total_harsh"]),
        ("Brusques moy. /100km", k["avg_harsh_100km"] if k["avg_harsh_100km"] else "n/a"),
        ("Véhicules à risque (Poor/Risk)", k["n_risk_poor"]),
        ("Conso suspecte (capteur)", k["n_conso_suspecte"]),
    ]
    _header_row(ws, 5, rows[0])
    for i, (lbl, val) in enumerate(rows[1:], start=6):
        a = ws.cell(row=i, column=1, value=lbl); b = ws.cell(row=i, column=2, value=val)
        a.font = Font(name="Arial", size=10); b.font = Font(name="Arial", size=10, bold=True, color=NAVY)
        a.border = BORDER; b.border = BORDER
        fill = LIGHT if i % 2 else WHITE
        a.fill = PatternFill("solid", fgColor=fill); b.fill = PatternFill("solid", fgColor=fill)
        b.alignment = Alignment(horizontal="right")
    _autosize(ws, {"A": 36, "B": 16})

    # ---- Scorecard combiné ----
    wss = wb.create_sheet("Scorecard combiné")
    sc = result["scorecard"]
    _title_band(wss, meta, k, ncols=len(sc.columns), title="SCORECARD FLOTTE")
    _add_logos(wss, len(sc.columns))
    hr = 5
    last = _write_table(wss, sc, hr, hr + 1)
    # ligne TOTAL/MOYENNE en formules
    tot = last + 1
    colmap = {c: idx + 1 for idx, c in enumerate(sc.columns)}
    def L(n): return get_column_letter(colmap[n])
    wss.cell(row=tot, column=1, value="TOTAL / MOYENNE")
    formulas = {
        "Distance (km)": f"=ROUND(SUM({L('Distance (km)')}{hr+1}:{L('Distance (km)')}{last}),0)",
        "Score (%)": f"=ROUND(AVERAGE({L('Score (%)')}{hr+1}:{L('Score (%)')}{last}),1)",
        "Conso (L/100km)": f"=ROUND(AVERAGE({L('Conso (L/100km)')}{hr+1}:{L('Conso (L/100km)')}{last}),1)",
        "Carburant (L)": f"=ROUND(SUM({L('Carburant (L)')}{hr+1}:{L('Carburant (L)')}{last}),0)",
        "Évén. brusques": f"=SUM({L('Évén. brusques')}{hr+1}:{L('Évén. brusques')}{last})",
    }
    for name, f in formulas.items():
        wss.cell(row=tot, column=colmap[name], value=f)
    for j in range(1, len(sc.columns) + 1):
        c = wss.cell(row=tot, column=j)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
    _autosize(wss, {"A": 30, "B": 12, "C": 10, "D": 14, "E": 15, "F": 12,
                    "G": 14, "H": 14, "I": 12, "J": 12})
    wss.freeze_panes = f"A{hr+1}"

    # ---- Comportement détail (par conducteur) ----
    wsb = wb.create_sheet("Comportement détail")
    bd = result["behavior_detail"]
    _title_band(wsb, meta, k, ncols=len(bd.columns), title="DÉTAIL COMPORTEMENT")
    _add_logos(wsb, len(bd.columns))
    _write_table(wsb, bd, 5, 6)
    _autosize(wsb, {"A": 28, "B": 22, "C": 12, "D": 14, "E": 16, "F": 14, "G": 13, "H": 12, "I": 9, "J": 11})
    wsb.freeze_panes = "A6"

    # ---- Consommation détail (par véhicule) ----
    wsc = wb.create_sheet("Consommation détail")
    cd = result["conso_detail"]
    _title_band(wsc, meta, k, ncols=len(cd.columns), title="DÉTAIL CONSOMMATION")
    _add_logos(wsc, len(cd.columns))
    _write_table(wsc, cd, 5, 6)
    _autosize(wsc, {"A": 28, "B": 14, "C": 14, "D": 13})
    wsc.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
